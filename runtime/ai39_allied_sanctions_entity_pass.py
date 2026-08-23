from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import tempfile
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Mapping

import requests
from openpyxl import load_workbook

SCHEMA = "AI39_ALLIED_SANCTIONS_ENTITY_V0_1"
TERMS = [
    "State Unitary Enterprise of the Donetsk People's Republic Republican Telecommunications Operator",
    "Republican Telecommunications Operator",
    "KOMTEL DPR",
    "KOMTEL",
    "KOMTEL-DPR-AS",
    "ORG-SUEO4-RIPE",
    "Timer LLC",
    "Respublikanskiy Operator Svyazi",
    "Respublikansky Operator Svyazi",
    "Республиканский оператор связи",
    "Государственное унитарное предприятие Донецкой Народной Республики Республиканский оператор связи",
]

SOURCES = {
    "uk_fcdo": "https://sanctionslist.fcdo.gov.uk/docs/UK-Sanctions-List.csv",
    "ofac_sdn": "https://www.treasury.gov/ofac/downloads/sdn.csv",
    "ofac_sdn_alt": "https://www.treasury.gov/ofac/downloads/alt.csv",
    "ofac_consolidated": "https://www.treasury.gov/ofac/downloads/cons_prim.csv",
    "ofac_consolidated_alt": "https://www.treasury.gov/ofac/downloads/cons_alt.csv",
    "canada_xml": "https://www.international.gc.ca/world-monde/assets/office_docs/international_relations-relations_internationales/sanctions/sema-lmes.xml",
    "australia_xlsx": "https://www.dfat.gov.au/sites/default/files/Australian_Sanctions_Consolidated_List.xlsx",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).casefold()
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("’", "'").replace("`", "'")
    text = re.sub(r"[^\w\u0400-\u04ff]+", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip()


def tokens(value: str) -> set[str]:
    stop = {
        "the", "of", "and", "state", "unitary", "enterprise", "llc", "ltd", "limited",
        "company", "joint", "stock", "public", "private", "organization", "org",
    }
    return {token for token in normalize(value).split() if len(token) >= 3 and token not in stop}


def score(term: str, candidate: str) -> tuple[float, str]:
    nt = normalize(term)
    nc = normalize(candidate)
    if not nt or not nc:
        return 0.0, "EMPTY"
    if nt in nc or nc in nt:
        shorter = min(len(nt), len(nc))
        if shorter >= 6:
            return 1.0, "SUBSTRING"
    tt, tc = tokens(nt), tokens(nc)
    if not tt or not tc:
        return 0.0, "NO_TOKENS"
    jaccard = len(tt & tc) / max(len(tt | tc), 1)
    containment = len(tt & tc) / max(min(len(tt), len(tc)), 1)
    sequence = SequenceMatcher(None, nt, nc).ratio()
    combined = max(jaccard, 0.75 * containment + 0.25 * sequence)
    return round(combined, 4), "FUZZY"


def match_record(source: str, record_id: str, text: str, fields: Mapping[str, Any] | None = None) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for term in TERMS:
        match_score, method = score(term, text)
        threshold = 0.82
        if normalize(term) in {"komtel", "timer llc"}:
            threshold = 0.96
        if match_score >= threshold:
            output.append(
                {
                    "source": source,
                    "record_id": record_id,
                    "search_term": term,
                    "match_score": match_score,
                    "match_method": method,
                    "record_excerpt": re.sub(r"\s+", " ", text).strip()[:1000],
                    "safe_fields": dict(fields or {}),
                }
            )
    return output


def get(url: str, timeout: float = 90) -> requests.Response:
    headers = {"User-Agent": "AI39-defensive-entity-research/0.1"}
    response = requests.get(url, timeout=timeout, headers=headers, allow_redirects=True)
    response.raise_for_status()
    return response


def parse_uk(content: bytes) -> tuple[int, list[dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    matches: list[dict[str, Any]] = []
    count = 0
    for row in reader:
        count += 1
        joined = " | ".join(str(value) for value in row.values() if value)
        record_id = str(row.get("Unique ID") or row.get("UK Sanctions List Ref") or count)
        safe = {
            "name": row.get("Name 1") or row.get("Name") or row.get("Primary Name"),
            "type": row.get("Designation Type") or row.get("Type"),
            "regime": row.get("Regime Name") or row.get("Regime"),
        }
        matches.extend(match_record("uk_fcdo", record_id, joined, safe))
    return count, matches


def parse_ofac_primary(source: str, content: bytes) -> tuple[int, list[dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    matches: list[dict[str, Any]] = []
    count = 0
    for row in csv.reader(io.StringIO(text)):
        if not row:
            continue
        count += 1
        record_id = row[0].strip() if row else str(count)
        name = row[1].strip() if len(row) > 1 else ""
        joined = " | ".join(row)
        safe = {
            "name": name,
            "type": row[2].strip() if len(row) > 2 else None,
            "program": row[3].strip() if len(row) > 3 else None,
        }
        matches.extend(match_record(source, record_id, joined, safe))
    return count, matches


def parse_ofac_alt(source: str, content: bytes) -> tuple[int, list[dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    matches: list[dict[str, Any]] = []
    count = 0
    for row in csv.reader(io.StringIO(text)):
        if not row:
            continue
        count += 1
        record_id = row[0].strip() if row else str(count)
        alias = row[3].strip() if len(row) > 3 else ""
        safe = {"alias": alias, "alias_type": row[2].strip() if len(row) > 2 else None}
        matches.extend(match_record(source, record_id, " | ".join(row), safe))
    return count, matches


def strip_tag(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def parse_canada(content: bytes) -> tuple[int, list[dict[str, Any]]]:
    root = ET.fromstring(content)
    children = list(root)
    if len(children) == 1 and len(list(children[0])) > 10:
        children = list(children[0])
    matches: list[dict[str, Any]] = []
    count = 0
    for element in children:
        fields: dict[str, str] = {}
        for node in element.iter():
            value = (node.text or "").strip()
            if value:
                key = strip_tag(node.tag)
                fields[key] = f"{fields[key]} | {value}" if key in fields else value
        if not fields:
            continue
        count += 1
        joined = " | ".join(fields.values())
        record_id = fields.get("Item") or fields.get("item") or fields.get("id") or str(count)
        safe = {
            "entity": fields.get("Entity") or fields.get("entity") or fields.get("Name") or fields.get("name"),
            "regulation": fields.get("Regulation") or fields.get("regulation"),
            "aliases": fields.get("Aliases") or fields.get("aliases"),
        }
        matches.extend(match_record("canada", str(record_id), joined, safe))
    return count, matches


def parse_australia(content: bytes) -> tuple[int, list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    matches: list[dict[str, Any]] = []
    count = 0
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header: list[str] | None = None
        for raw in rows:
            values = ["" if value is None else str(value).strip() for value in raw]
            nonempty = [value for value in values if value]
            if not nonempty:
                continue
            if header is None:
                normalized = [normalize(value) for value in values]
                if any("name" in value for value in normalized) and any(
                    "reference" in value or value == "ref" for value in normalized
                ):
                    header = values
                continue
            count += 1
            row = {header[index] if index < len(header) and header[index] else f"col_{index}": value for index, value in enumerate(values) if value}
            joined = " | ".join(row.values())
            record_id = next((value for key, value in row.items() if "reference" in normalize(key)), str(count))
            safe = {
                "name": next((value for key, value in row.items() if "name" in normalize(key)), None),
                "type": next((value for key, value in row.items() if normalize(key) == "type"), None),
                "regime": next((value for key, value in row.items() if "regime" in normalize(key)), None),
            }
            matches.extend(match_record("australia", str(record_id), joined, safe))
    return count, matches


def parse_ripe() -> dict[str, Any]:
    response = get("https://stat.ripe.net/data/whois/data.json?resource=AS202279", timeout=45)
    payload = response.json()
    records = payload.get("data", {}).get("records", []) if isinstance(payload, Mapping) else []
    allowed = {"aut-num", "as-name", "descr", "org", "organisation", "org-name", "org-type", "country", "mnt-by", "mnt-ref", "created", "last-modified", "source", "status"}
    extracted: list[dict[str, str]] = []
    for group in records if isinstance(records, list) else []:
        if not isinstance(group, list):
            continue
        item: dict[str, str] = {}
        for field in group:
            if not isinstance(field, Mapping):
                continue
            key = str(field.get("key") or "").casefold()
            value = str(field.get("value") or "").strip()
            if key in allowed and value:
                item[key] = value
        if item:
            extracted.append(item)
    return {
        "source": "RIPEstat whois",
        "resource": "AS202279",
        "records": extracted,
        "response_sha256": sha256_bytes(response.content),
    }


def dedupe(matches: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    output: list[dict[str, Any]] = []
    for match in matches:
        key = hashlib.sha256(
            json.dumps(
                {
                    "source": match.get("source"),
                    "record_id": match.get("record_id"),
                    "search_term": match.get("search_term"),
                    "excerpt": match.get("record_excerpt"),
                },
                sort_keys=True,
                ensure_ascii=False,
            ).encode("utf-8")
        ).hexdigest()
        if key in seen:
            continue
        seen.add(key)
        output.append(dict(match))
    output.sort(key=lambda item: (float(item.get("match_score", 0.0)), str(item.get("source"))), reverse=True)
    return output


def atomic_write(path: str | Path, value: Mapping[str, Any]) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=target.parent, delete=False) as handle:
        handle.write(payload)
        handle.flush()
        os.fsync(handle.fileno())
        temp = Path(handle.name)
    os.replace(temp, target)


def run(out_path: str | Path) -> None:
    parsers = {
        "uk_fcdo": parse_uk,
        "ofac_sdn": lambda content: parse_ofac_primary("ofac_sdn", content),
        "ofac_sdn_alt": lambda content: parse_ofac_alt("ofac_sdn_alt", content),
        "ofac_consolidated": lambda content: parse_ofac_primary("ofac_consolidated", content),
        "ofac_consolidated_alt": lambda content: parse_ofac_alt("ofac_consolidated_alt", content),
        "canada_xml": parse_canada,
        "australia_xlsx": parse_australia,
    }
    source_state: dict[str, Any] = {}
    all_matches: list[dict[str, Any]] = []
    for source, url in SOURCES.items():
        try:
            response = get(url)
            row_count, matches = parsers[source](response.content)
            source_state[source] = {
                "status": "OK",
                "final_url": response.url,
                "http_status": response.status_code,
                "bytes": len(response.content),
                "sha256": sha256_bytes(response.content),
                "record_count": row_count,
                "match_count": len(matches),
            }
            all_matches.extend(matches)
        except Exception as exc:
            source_state[source] = {
                "status": "DEGRADED",
                "error_type": type(exc).__name__,
                "error": str(exc)[:500],
            }

    try:
        ripe = parse_ripe()
        ripe_status = "OK"
    except Exception as exc:
        ripe = {"error_type": type(exc).__name__, "error": str(exc)[:500]}
        ripe_status = "DEGRADED"

    unique_matches = dedupe(all_matches)
    exact = [match for match in unique_matches if match.get("match_method") == "SUBSTRING"]
    fuzzy = [match for match in unique_matches if match.get("match_method") == "FUZZY"]
    if exact:
        state = "SUPPORTED_DIRECT_LIST_MATCH"
        confidence = 0.92
    elif fuzzy:
        state = "UNRESOLVED_CANDIDATE_MATCHES"
        confidence = 0.45
    else:
        state = "NO_DIRECT_MATCH_ACROSS_AVAILABLE_LISTS"
        confidence = 0.30

    result = {
        "schema_version": SCHEMA,
        "generated_at": utc_now(),
        "scope": "Official allied sanctions lists and RIPE organization metadata; entity-level only, no hosts, IPs, credentials, vulnerabilities or exact operational locations.",
        "search_terms": TERMS,
        "source_state": source_state,
        "ripe_status": ripe_status,
        "ripe": ripe,
        "matches": unique_matches[:100],
        "direct_match_count": len(exact),
        "fuzzy_match_count": len(fuzzy),
        "candidate_state": state,
        "confidence": confidence,
        "interpretation_rules": [
            "A list match establishes only that the named record appears in that official sanctions dataset; it does not prove current operational activity or physical location.",
            "No direct match is weak negative evidence only and never establishes legitimacy, ownership or non-sanctioned control.",
            "Fuzzy matches remain leads until the exact record and alias relationship are manually verified.",
            "Sanctions datasets may differ by jurisdiction, update time, ownership rules and transliteration.",
        ],
        "next_gap": "If a direct or high-quality alias match exists, expand only through official designation notices, ownership/control guidance and counterparties. Otherwise prioritize procurement/integrator and RF/VHR evidence rather than repeating list searches.",
        "result_sha256": "",
    }
    material = dict(result)
    material["result_sha256"] = ""
    result["result_sha256"] = hashlib.sha256(
        json.dumps(material, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    atomic_write(out_path, result)
    print(
        json.dumps(
            {
                "candidate_state": result["candidate_state"],
                "direct_match_count": result["direct_match_count"],
                "fuzzy_match_count": result["fuzzy_match_count"],
                "sources": {key: value.get("status") for key, value in source_state.items()},
                "top_matches": unique_matches[:10],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="AI-39 allied sanctions and entity-neighbor pass")
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    run(args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
